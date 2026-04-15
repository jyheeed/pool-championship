#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "player"


def to_int(value) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_bd_joueurs.py <path-to-BD_joueurs.xlsx>")
        return 1

    input_path = Path(sys.argv[1]).resolve()
    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        return 1

    wb = load_workbook(input_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Workbook is empty.")
        return 1

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]

    output_dir = input_path.parent / "import_output"
    output_dir.mkdir(exist_ok=True)

    mapped_players = []

    for index, row in enumerate(data_rows, start=1):
        if not any(cell not in (None, "") for cell in row):
            continue

        values = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        # Support both raw French source columns and already-normalized player schema.
        first_name = str(values.get("Prénom") or values.get("Prenom") or "").strip()
        last_name = str(values.get("Nom") or "").strip()
        normalized_name = str(values.get("name") or "").strip()
        full_name = " ".join(part for part in [first_name, last_name] if part).strip() or normalized_name or f"Player {index}"
        direct_id = str(values.get("id") or "").strip()
        direct_nationality = str(values.get("nationality") or "").strip()
        direct_age = str(values.get("age") or "").strip()
        direct_nickname = str(values.get("nickname") or "").strip()
        direct_photo = str(values.get("photo_url") or "").strip()
        direct_group = str(values.get("pool_group") or "").strip()
        direct_club = str(values.get("club") or values.get("Club") or "").strip()

        mapped_players.append(
            {
                "id": direct_id or f"tn-{slugify(full_name)}",
                "name": full_name,
                "nickname": direct_nickname,
                "nationality": direct_nationality or "Tunisia",
                "age": direct_age,
                "club": direct_club,
                "photo_url": direct_photo,
                "pool_group": direct_group,
                "source_points": to_int(values.get("Points")),
                "source_rank": to_int(values.get("Rang")),
                "source_wins": to_int(values.get("Win")),
                "source_losses": to_int(values.get("Loose")),
                "source_games": to_int(values.get("Games")),
            }
        )

    if not mapped_players:
        print("No player rows were found under the header row.")
        print(f"Headers detected: {headers}")
        return 0

    mongo_csv_rows = [["id", "name", "nickname", "nationality", "age", "club", "photo_url", "pool_group"]]
    for player in mapped_players:
        mongo_csv_rows.append([
            player["id"],
            player["name"],
            player["nickname"],
            player["nationality"],
            player["age"],
            player["club"],
            player["photo_url"],
            player["pool_group"],
        ])

    seed_json_path = output_dir / "players.seed.json"
    seed_json_path.write_text(json.dumps(mapped_players, indent=2, ensure_ascii=False), encoding="utf-8")

    csv_path = output_dir / "players.mongo.csv"
    csv_lines = [",".join(row) for row in mongo_csv_rows]
    csv_path.write_text("\n".join(csv_lines), encoding="utf-8")

    report_path = output_dir / "import-summary.txt"
    report_path.write_text(
        "\n".join(
            [
                f"Input file: {input_path.name}",
                f"Sheet used: {ws.title}",
                f"Headers detected: {', '.join(headers)}",
                f"Players mapped: {len(mapped_players)}",
                "",
                "Output files:",
                f"- {seed_json_path.name}",
                f"- {csv_path.name}",
                "",
                "Next step:",
                "Use the JSON file with npm run seed:players or import the CSV into Mongo tooling.",
            ]
        ),
        encoding="utf-8",
    )

    print(f"Import complete. Output folder: {output_dir}")
    print(f"Players mapped: {len(mapped_players)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
